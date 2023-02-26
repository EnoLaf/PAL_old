from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import json

def excel_to_json():
    wb = load_workbook(filename='D:/Code/PAL/pal2.xlsx')
    ws = wb.active

    my_list = []

    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))

    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        my_list.append(my_dict)

    del my_list[0]
    data = json.dumps(my_list, sort_keys=True, indent=4)
    with open('D:/Code/PAL/pal.json', 'w', encoding='utf-8') as f:
        f.write(data)
