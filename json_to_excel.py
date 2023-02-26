import json
from openpyxl import Workbook

def json_to_excel():
    keys = []
    wb = Workbook()
    ws = wb.active

    with open('D:/Code/PAL/pal.json',encoding="utf8") as f:
        json_data = json.load(f)

    for i in range(len(json_data)) :
    	sub_obj = json_data[i]
    	if i == 0 :
    		keys = list(sub_obj.keys())
    		for k in range(len(keys)) :
    			# row or column index start from 1
    			ws.cell(row = (i + 1), column = (k + 1), value = keys[k]);
    	for j in range(len(keys)) :
    		ws.cell(row = (i + 2), column = (j + 1), value = sub_obj[keys[j]]);
    wb.save("D:/Code/PAL/pal2.xlsx")
