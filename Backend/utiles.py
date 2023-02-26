import random
import json
import tkinter
import tkinter.font as tkFont
from PIL            import Image, ImageTk
from tkinter        import ttk
from openpyxl       import Workbook
from openpyxl       import load_workbook
from openpyxl.utils import get_column_letter
import sys

sys.path.append("D:\Code\PAL")
from Frontend.utiles import Functions

F = Functions()

class Sortbooks():

    def __init__(self):
        self.min = 0
        self.max = 0

    def excel_to_json(self):
        """ Convert data base from excel to json

                Parameters
                ----------
                excel_path (str)
                    path of excel file
                json_path (str)
                    path of json file

                Returns
                -------
                /

            """
        wb = load_workbook(filename='D:/Code/PAL/Backend/pal.xlsx')
        ws = wb.active

        my_list = []

        last_column = len(list(ws.columns))
        last_row = len(list(ws.rows))

        for row in range(1, last_row + 1):
            my_dict = {}
            for column in range(1, last_column + 1):
                column_letter = get_column_letter(column)
                if row > 1:
                    my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
            my_list.append(my_dict)

        del my_list[0]
        data = json.dumps(my_list, sort_keys=True, indent=4)
        with open('D:/Code/PAL/Backend/pal.json', 'w', encoding='utf-8') as f:
            f.write(data)


    def json_to_excel(self):
        """ Convert data base from json to excel

                Parameters
                ----------
                json_path (str)
                    path of json file
                excel_path (str)
                    path of excel file

                Returns
                -------
                /

            """
        keys = []
        wb = Workbook()
        ws = wb.active

        with open('D:/Code/PAL/Backend/pal.json',encoding="utf8") as f:
            json_data = json.load(f)

        for i in range(len(json_data)) :
        	sub_obj = json_data[i]
        	if i == 0 :
        		keys = list(sub_obj.keys())
        		for k in range(len(keys)) :
        			# row or column index start from 1
        			ws.cell(row = (i + 1), column = (k + 1), value = keys[k]);
        	for j in range(len(keys)) :
        		ws.cell(row = (i + 2), column = (j + 1), value = sub_obj[keys[j]]);
        wb.save('D:/Code/PAL/Backend/pal.xlsx')


    def parse_pages(self, pages):
        """ Scroll through the list of pages of the registered books to extract
                the indexes that meet the requirements

                Parameters
                ----------
                pages (list)
                    List of all number of pages of the registered books

                Returns
                -------
                index_list (list)
                     List of indexes in the book list that match the filters
            """
        index_list = []
        for i in range(len(pages)):
            if self.max is None:
                if pages[i]>self.min:
                    index_list.append(i)
            elif pages[i]>self.min and pages[i]<self.max:
                index_list.append(i)
        return index_list


    def search_book(self, books, index_list):
        """ Uses the index list to retrieve data about the corresponding
                to the corresponding books

                Parameters
                ----------
                books (list)
                    List of books in the DB
                index_list (list)
                    List of indexes in the book list that match the filters

                Returns
                -------
                list_size_size_filtered_books (list)
                    List of books that match the filters
            """
        list_size_size_filtered_books = []
        for i in range(len(books)):
            if i in index_list:
                list_size_size_filtered_books.append(books[i])
        return list_size_size_filtered_books

    def size_filtering(self, list_nb_pages, data):
        """ Filter the books in the DB that match the page requirements

                Parameters
                ----------
                list_nb_pages (list)
                    List of books in the DB
                data (list)
                    List of books in the DB

                Returns
                -------
                size_filtered_books

            """
        index_page          = self.parse_pages(list_nb_pages)
        size_filtered_books = self.search_book(data, index_page)
        return size_filtered_books

    def user_input(self):
        """ Asks the user for the size of the book and its style

                Parameters
                ----------
                None
                    ...

                Returns
                -------
                size (int)
                    Number of pages requested by the user
                style (str)
                    Style requested by the user
            """
        size=int(input("Veux-tu lire un livre : \n" "\t1-Petit <400 pages\n"
                                                    "\t2-Moyen 400 à 600 pages\n"
                                                    "\t3-Gros 600 à 800 pages\n"
                                                    "\t4-Très gros >800 pages\n"
                      )
                )

        style=input("Quel genre de livre veux-tu?\n")
        return size, style

    def style_filtering(self, style, size_filtered_books):
        """ Filter the books (already filtered by the size requirements)
                by the style requirements

                Parameters
                ----------
                style (str)
                    Style requested by the user
                size_filtered_books (list)

                    ...

                Returns
                -------
                list_style_filtered_books (list)
                    list of books corresponding to the requirements of size and style
            """
        i=0
        list_style_filtered_books = []
        while i<len(size_filtered_books):
            if style in size_filtered_books[i]["style"]:
                list_style_filtered_books.append(size_filtered_books[i])
            i+=1
        return list_style_filtered_books


    def random_filtered_book(self, list_style_filtered_books):
        """ Randomly choose a book among those corresponding to
                the requirements

                Parameters
                ----------
                list_style_filtered_books (list)
                    list of books corresponding to the requirements of size and style

                    ...

                Returns
                -------
                chosen_book_title (str)
                    title of the book chosen randomly among the requirements
                chosen_book_subtitle (str)
                    subtitle of the book chosen randomly among the requirements
                chosen_book_writer (str)
                    author of the book chosen randomly among the requirements
                chosen_book_cover (str)
                    cover of the book chosen randomly among the requirements
                chosen_book_style (str)
                    styles of the book chosen randomly among the requirements

            """
        if not list_style_filtered_books :
            chosen_book_title       = "Aucun livre ne correspond"
            chosen_book_subtitle    = "Réessayez avec d'autres critères"
            chosen_book_writer      = ""
            chosen_book_cover       = "404"
            chosen_book_style       = ""
        else:
            random_int = random.randint(0,len(list_style_filtered_books)-1)
            chosen_book_title       = list_style_filtered_books[random_int]["title"]
            chosen_book_subtitle    = list_style_filtered_books[random_int]["subtitle"]
            chosen_book_writer      = list_style_filtered_books[random_int]["writer"]
            chosen_book_cover       = list_style_filtered_books[random_int]["cover"]
            chosen_book_style       = list_style_filtered_books[random_int]["style"]

        return chosen_book_title, chosen_book_subtitle, chosen_book_writer, chosen_book_cover, chosen_book_style
